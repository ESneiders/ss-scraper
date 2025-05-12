import os
import time
import openpyxl

class Sludinajums:
    def __init__(self, url, cena, nosaukums, apraksts):
        self.url = url
        self.cena = cena
        self.nosaukums = nosaukums
        self.apraksts = apraksts
        self.status = None  # Statuss: "Jauns", "None", "Izņemts"

    def __str__(self):
        return f"{self.nosaukums}: {self.cena} EUR - {self.url}"

class Kategorija:
    def __init__(self, name, url):
        self.name = name
        self.url = url

    def get_url(self):
        return self.url

class DatuApstrade:
    def __init__(self, kategorija, file_name):
        self.kategorija = kategorija
        self.file_name = file_name
        self.sludinajumi = []
        self.old_urls = set()
        self.load_old_data()

    def load_old_data(self):
        """ Ielādē esošos datus no Excel faila (vai citādi saglabātajiem datiem) """
        if os.path.exists(self.file_name):
            wb = openpyxl.load_workbook(self.file_name)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):  # sāk no 2. rindas
                url = row[0]
                if url:
                    self.old_urls.add(url)
            wb.close()

    def save_data(self):
        """ Saglabā datus Excel failā """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sludinājumi"
        
        # Piešķir kolonnu nosaukumus
        ws.append(["SAITE", "Cena", "Nosaukums", "Apraksts", "Status"])
        
        for sludinajums in self.sludinajumi:
            ws.append([sludinajums.url, sludinajums.cena, sludinajums.nosaukums, sludinajums.apraksts, sludinajums.status])
        
        wb.save(self.file_name)  # Saglabā failu ar norādīto nosaukumu
        print(f"Dati saglabāti kā: {self.file_name}")

    def __init__(self, kategorija, file_name):
        self.kategorija = kategorija
        self.file_name = file_name
        self.sludinajumi = []
        self.old_urls = set()
        self.load_old_data()

    def load_old_data(self):
        """ Ielādē esošos datus no Excel faila (vai citādi saglabātajiem datiem) """
        if os.path.exists(self.file_name):
            with open(self.file_name, 'r') as f:
                lines = f.readlines()
                for line in lines[1:]:  # Pārskaitām tikai datus pēc virsraksta
                    data = line.strip().split(',')
                    url = data[0]
                    if url:
                        self.old_urls.add(url)

    def save_data(self):
        """ Saglabā datus Excel failā (vienkāršota versija - CSV) """
        with open(self.file_name, 'w', encoding='utf-8') as f:
            f.write("SAITE," + "Cena,Nosaukums,Apraksts,Status\n")
            for sludinajums in self.sludinajumi:
                f.write(f"{sludinajums.url},{sludinajums.cena},{sludinajums.nosaukums},{sludinajums.apraksts},{sludinajums.status}\n")
        print(f"Dati saglabāti kā: {self.file_name}")

    def check_new_and_removed(self):
        """ Pārbauda jaunus un izņemtus sludinājumus """
        new_urls = set([s.url for s in self.sludinajumi])
        for sludinajums in self.sludinajumi:
            if sludinajums.url in self.old_urls:
                sludinajums.status = "None"
            else:
                sludinajums.status = "Jauns"
        
        # Pārbauda vecos datus, kas vairs nav pieejami
        removed_sludinajumi = [sl for sl in self.sludinajumi if sl.url not in new_urls]
        for sl in removed_sludinajumi:
            sl.status = "Izņemts"

    def scrape_data(self, pages=1):
        """ Šī funkcija parāda imitāciju sludinājumu ievākšanai no tīmekļa """
        for page in range(1, pages + 1):
            url = self.kategorija.get_url() + f"/page{page}.html" if page > 1 else self.kategorija.get_url()
            print(f"Skata lapu: {url}")
            time.sleep(1)  # Imitē ielādi

            # Imitē datus no lapas
            for i in range(5):  # Imitē 5 sludinājumus katrā lapā
                sludinajums = Sludinajums(
                    url=f"{url}/sludinajums_{i}",
                    cena=100 + i * 10,  # Cena piemērs
                    nosaukums=f"Sludinājums {i}",
                    apraksts=f"Apraksts {i} par produktu"
                )
                self.sludinajumi.append(sludinajums)

        self.check_new_and_removed()

def main():
    # Izvēle no kategorijām
    kategorijas = [
        Kategorija("Auto (Alfa Romeo)", "https://www.ss.lv/lv/transport/cars/alfa-romeo/"),
        Kategorija("Darbs (Administrators)", "https://www.ss.lv/lv/home-stuff/furniture/"),
        Kategorija("Suņi (viss)", "https://www.ss.lv/lv/animals/dogs/"),
    ]
    
    print("Izvēlieties kategoriju:")
    for idx, cat in enumerate(kategorijas, 1):
        print(f"{idx}. {cat.name}")

    choice = int(input("Ievadiet numuru: ")) - 1
    selected_cat = kategorijas[choice]

    file_name = input("Ievadiet Excel faila nosaukumu (piem. dati.csv): ").strip()
    datu_apstrade = DatuApstrade(selected_cat, file_name)

    pages = int(input("Cik lapas pārbaudīt? "))
    datu_apstrade.scrape_data(pages)
    
    datu_apstrade.save_data()

if __name__ == "__main__":
    main()
